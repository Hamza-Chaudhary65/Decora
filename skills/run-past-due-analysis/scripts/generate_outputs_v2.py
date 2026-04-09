import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
import json

TODAY = date(2026, 4, 2)
results_df = pd.read_pickle('/home/claude/results_df_v2.pkl')

header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
header_fill = PatternFill('solid', start_color='1F4E79')
tier_colors = {1: 'C00000', 2: 'FF9900', 3: '375623', 4: '7F7F7F'}

def write_sheet(ws, df, col_widths, tier_col=None):
    for c, (h, w) in enumerate(zip(df.columns, col_widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    for r, row in enumerate(df.itertuples(index=False), 2):
        tier = row[0] if tier_col else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        if tier_col and tier in tier_colors:
            tc = ws.cell(r, 2)
            tc.fill = PatternFill('solid', start_color=tier_colors[tier])
            tc.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        elif r % 2 == 0:
            for c in range(1, len(df.columns)+1):
                if ws.cell(r, c).fill.start_color.rgb in ('00000000','FFFFFFFF'):
                    ws.cell(r, c).fill = PatternFill('solid', start_color='EBF3FA')
    ws.freeze_panes = 'A2'

# FILE 1: NetSuite Upload
upload_df = results_df[['Internal ID','Sales Order','Item Upload Name','Item','Existing Procurement Notes','SC Procurement Update']].copy()
wb1 = Workbook(); ws1 = wb1.active; ws1.title = 'SC Procurement Update'
write_sheet(ws1, upload_df, [15, 15, 18, 30, 50, 95])
wb1.save('/home/claude/SC_Updates_NetSuite_Upload_FINAL.xlsx')
print(f"File 1: {len(upload_df)} rows")

# FILE 2: Full Detail
detail_cols = ['Priority_Tier','Tier_Label','Internal ID','Sales Order','Item Upload Name','Item',
               'Company Name','Backorder Qty','Original Promise Date','Is Past Due',
               'Allocated PO','PO Exp Ship Date','Is In Transit',
               'Existing Procurement Notes','CSR Request Notes','SC Procurement Update']
detail_df = results_df[detail_cols].copy()
wb2 = Workbook(); ws2 = wb2.active; ws2.title = 'Full Detail'
write_sheet(ws2, detail_df, [8,12,15,15,16,28,25,10,18,10,12,14,10,40,40,90], tier_col=True)
ws2.auto_filter.ref = f'A1:{get_column_letter(len(detail_cols))}1'
wb2.save('/home/claude/SC_Procurement_Updates_Upload_FINAL.xlsx')
print(f"File 2: {len(detail_df)} rows")

# FILE 3: Daily Follow-Up Target List
# Include a line only when ALL THREE conditions are true:
#   1. Past due â€” Original Promise Date is before today
#   2. Not in transit â€” material has not shipped within the last 14 days
#   3. No future warehouse availability date â€” comment contains no WH avail date after today
import re as _re

def extract_wh_avail(comment):
    matches = _re.findall(r'Expected warehouse availability (\d{2}/\d{2}/\d{4})', str(comment))
    if matches:
        try:
            return pd.to_datetime(matches[-1]).date()
        except:
            return None
    return None

results_df['wh_avail_date'] = results_df['SC Procurement Update'].apply(extract_wh_avail)

followup_df = results_df[
    (results_df['Is Past Due']) &
    (~results_df['Is In Transit']) &
    (results_df['wh_avail_date'].isna() | (results_df['wh_avail_date'] <= TODAY))
].copy()

followup_df['WH Avail Date'] = followup_df['wh_avail_date'].apply(
    lambda d: d.strftime('%m/%d/%Y') if d else ''
)

fu_cols = ['Priority_Tier','Tier_Label','Sales Order','Item','Company Name','Backorder Qty',
           'Original Promise Date','Allocated PO','PO Exp Ship Date','WH Avail Date',
           'Existing Procurement Notes','CSR Request Notes','SC Procurement Update']
followup_df = followup_df[fu_cols].sort_values(['Priority_Tier','Original Promise Date'])
wb3 = Workbook(); ws3 = wb3.active; ws3.title = 'Daily Follow-Up Targets'
ws3.sheet_properties.tabColor = '7030A0'
for c, (h, w) in enumerate(zip(followup_df.columns, [8,12,15,28,25,10,18,12,14,14,40,40,90]), 1):
    cell = ws3.cell(1, c, h)
    cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.fill = PatternFill('solid', start_color='7030A0')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws3.column_dimensions[get_column_letter(c)].width = w
ws3.row_dimensions[1].height = 30
fu_tier_colors = {2: 'FF9900', 3: '375623', 4: '7F7F7F'}
for r, row in enumerate(followup_df.itertuples(index=False), 2):
    tier = row[0]
    for c, val in enumerate(row, 1):
        cell = ws3.cell(r, c, val)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    tc = ws3.cell(r, 2)
    if tier in fu_tier_colors:
        tc.fill = PatternFill('solid', start_color=fu_tier_colors[tier])
        tc.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    if r % 2 == 0:
        for c in range(1, len(followup_df.columns)+1):
            if ws3.cell(r, c).fill.start_color.rgb in ('00000000', 'FFFFFFFF'):
                ws3.cell(r, c).fill = PatternFill('solid', start_color='F3EEFA')
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f'A1:{get_column_letter(len(followup_df.columns))}1'
wb3.save('/home/claude/Daily_FollowUp_Target_List.xlsx')
print(f"File 3: {len(followup_df)} follow-up targets")

# STATS
total = len(results_df)
covered = (results_df['Allocated PO'] != 'NONE').sum()
stats = {
    'total': total, 'covered': int(covered),
    'coverage_pct': round(covered/total*100, 1),
    'tier_counts': results_df['Tier_Label'].value_counts().to_dict(),
    'in_transit': int(results_df['Is In Transit'].sum()),
    'past_due': int(results_df['Is Past Due'].sum()),
    'followup_count': len(followup_df)
}
with open('/home/claude/stats.json','w') as f:
    json.dump(stats, f)
print(f"\nStats: {stats}")
